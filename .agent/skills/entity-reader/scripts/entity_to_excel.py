#!/usr/bin/env python3
"""
entity_to_excel.py
------------------
Generic helper: takes a list of entity dicts and writes a formatted Excel workbook.

Usage (called from Claude's generated code):
    from scripts.entity_to_excel import write_excel
    write_excel(entities, output_path="output.xlsx")

Entity format:
    [
      {
        "entity_name": "User",
        "table_name":  "users",
        "fields": [
          {
            "field_name": "id", "column_name": "id",
            "code_type": "Long", "sql_type": "BIGINT",
            "pk": True, "nullable": False, "unique": True,
            "length": "", "precision": "", "scale": "",
            "default": "", "relation": "", "validation": "", "description": ""
          }, ...
        ]
      }, ...
    ]
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL   = PatternFill("solid", start_color="4472C4")
ALT_FILL      = PatternFill("solid", start_color="EBF3FB")
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF")
PK_FILL       = PatternFill("solid", start_color="E2EFDA")   # light green
FK_FILL       = PatternFill("solid", start_color="FFF2CC")   # light yellow
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
NORMAL_FONT   = Font(name="Arial", size=10)
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

COLUMNS = [
    ("Field Name",   "field_name",   18),
    ("Column Name",  "column_name",  18),
    ("Code Type",    "code_type",    16),
    ("SQL Type",     "sql_type",     16),
    ("PK",           "pk",            6),
    ("Nullable",     "nullable",      9),
    ("Unique",       "unique",        8),
    ("FK / Relation","relation",     22),
    ("Length",       "length",        8),
    ("Precision",    "precision",     9),
    ("Scale",        "scale",         7),
    ("Default",      "default",      14),
    ("Validation",   "validation",   22),
    ("Description",  "description",  30),
]


def _write_entity_sheet(wb: Workbook, entity: dict) -> None:
    name = entity["entity_name"][:31]   # Excel sheet name limit
    ws = wb.create_sheet(title=name)

    # Sub-title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    cell = ws.cell(row=1, column=1,
                   value=f"{entity['entity_name']}  →  table: {entity['table_name']}")
    cell.font    = Font(bold=True, name="Arial", size=11, color="1F3864")
    cell.fill    = PatternFill("solid", start_color="D9E1F2")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Header row
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=2, column=col_idx, value=header)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 18

    ws.freeze_panes = "A3"

    for row_idx, field in enumerate(entity.get("fields", []), start=3):
        alt = (row_idx % 2 == 1)
        is_pk = bool(field.get("pk"))
        is_fk = bool(field.get("relation"))
        row_fill = PK_FILL if is_pk else (FK_FILL if is_fk else (ALT_FILL if alt else WHITE_FILL))

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            val = field.get(key, "")
            if isinstance(val, bool):
                val = "✓" if val else ""
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font      = NORMAL_FONT
            c.fill      = row_fill
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border    = THIN_BORDER


def write_excel(entities: list, output_path: str = "entity_dictionary.xlsx",
                combined_threshold: int = 5) -> str:
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # Summary sheet
    ws_sum = wb.create_sheet(title="Summary", index=0)
    sum_headers = ["Entity Name", "Table Name", "Field Count", "PK Field(s)"]
    for ci, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center"); c.border = THIN_BORDER
    widths = [24, 24, 12, 30]
    for ci, w in enumerate(widths, 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws_sum.freeze_panes = "A2"

    for ri, entity in enumerate(entities, start=2):
        pks = [f["field_name"] for f in entity.get("fields", []) if f.get("pk")]
        ws_sum.cell(row=ri, column=1, value=entity["entity_name"]).border = THIN_BORDER
        ws_sum.cell(row=ri, column=2, value=entity["table_name"]).border  = THIN_BORDER
        ws_sum.cell(row=ri, column=3, value=len(entity.get("fields", []))).border = THIN_BORDER
        ws_sum.cell(row=ri, column=4, value=", ".join(pks)).border = THIN_BORDER
        fill = ALT_FILL if ri % 2 == 1 else WHITE_FILL
        for ci in range(1, 5):
            ws_sum.cell(row=ri, column=ci).fill = fill
            ws_sum.cell(row=ri, column=ci).font = NORMAL_FONT

    if len(entities) <= combined_threshold:
        # All entities on one "All Fields" sheet
        ws_all = wb.create_sheet(title="All Fields")
        all_cols = [("Entity", "entity_name", 18), ("Table", "table_name", 18)] + COLUMNS
        for ci, (h, _, w) in enumerate(all_cols, 1):
            c = ws_all.cell(row=1, column=ci, value=h)
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center"); c.border = THIN_BORDER
            ws_all.column_dimensions[get_column_letter(ci)].width = w
        ws_all.freeze_panes = "A2"
        row_idx = 2
        for entity in entities:
            for field in entity.get("fields", []):
                alt = row_idx % 2 == 0
                is_pk = bool(field.get("pk"))
                is_fk = bool(field.get("relation"))
                row_fill = PK_FILL if is_pk else (FK_FILL if is_fk else (ALT_FILL if alt else WHITE_FILL))
                ws_all.cell(row=row_idx, column=1, value=entity["entity_name"]).fill = row_fill
                ws_all.cell(row=row_idx, column=2, value=entity["table_name"]).fill  = row_fill
                for ci, (_, key, _) in enumerate(COLUMNS, start=3):
                    val = field.get(key, "")
                    if isinstance(val, bool):
                        val = "✓" if val else ""
                    c = ws_all.cell(row=row_idx, column=ci, value=val)
                    c.fill = row_fill; c.border = THIN_BORDER; c.font = NORMAL_FONT
                    c.alignment = Alignment(wrap_text=True)
                row_idx += 1
    else:
        for entity in entities:
            _write_entity_sheet(wb, entity)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # Quick smoke test
    sample = [
        {
            "entity_name": "User", "table_name": "users",
            "fields": [
                {"field_name": "id", "column_name": "id", "code_type": "Long",
                 "sql_type": "BIGINT", "pk": True, "nullable": False, "unique": True,
                 "length": "", "precision": "", "scale": "", "default": "",
                 "relation": "", "validation": "@NotNull", "description": "Primary key"},
                {"field_name": "email", "column_name": "email", "code_type": "String",
                 "sql_type": "VARCHAR(255)", "pk": False, "nullable": False, "unique": True,
                 "length": "255", "precision": "", "scale": "", "default": "",
                 "relation": "", "validation": "@Email @NotBlank", "description": "User email address"},
            ]
        }
    ]
    path = write_excel(sample, "/tmp/test_entity.xlsx")
    print(f"Written: {path}")
